from tkinter import*
from datetime import date 
from tkinter import filedialog 
from tkinter import messagebox 
from PIL import Image, ImageTk 
import os 
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib 
background="#06383D"
framebg="#EDEDED"
framefg="#06283D"
root=Tk()
root.title("Dot Student Internship details")
root.geometry("1250x700+210+100")
root.config(bg=background) 


file=pathlib.Path('student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active 
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="State"
    sheet['I1']="Company Name"
    sheet['J1']="company address"
    sheet['K1']="Internship duration"

    
    file.save('student_data.xlsx')







#upload image
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename( title="select Image File",filetype=[("JPG ile","*.jpg"),("PNG File","*.png"),("all files","*txt")]) 
    img =(Image.open(filename))
    resized_image= img.resize((190 ,190))
    photo2= ImageTk.PhotoImage(resized_image)   
    lbl.config(image=photo2)
    lbl.image=photo2
    
    
    
#registration number    
def registration_no():
    file=openpyxl.load_workbook('student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    
    max_row__value =  sheet.cell(row=row,column=1).value
    try:
        Registration.set(max_row__value+1)
    except:
        Registration.set("1")
        
        
        
        
        
#search
def search():
    text=Search.get()
    Clear()
    saveButton.config(state='disable')
    
    file=openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            print(str(name))
            reg_no_positions=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    try:
         print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!!")
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
   
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4=="Female":
        R2.select()
    else:
        R1.select()
    
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    State.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    
    
    
    img =(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image= img.resize((190 ,190))
    photo2= ImageTk.PhotoImage(resized_image)   
    lbl.config(image=photo2)
    lbl.image=photo2
#update
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=State.get()
    father_name=F_Name.get()
    mother_name=M_Name.get()
    F1=Father_Occupation.get()
    
    file=openpyxl.load_workbook("student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            print(reg_number)
    sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Re1)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=father_name)
    sheet.cell(column=10,row=int(reg_number),value=mother_name)
    sheet.cell(column=11,row=int(reg_number),value=F1)

    file.save('student_data.xlsx')
    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update Successful!!!!")
    Clear()
    
        
        
        
#clear

def Clear():
    global img
    Name.set('')
    DOB.set('')
    State.set('')
    Religion.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')

    Class.set("Select Class")
    registration_no()
    saveButton.config(state= 'normal')
    
    img1=PhotoImage(file='Images/upload.png')
    lbl.config(image=img1)
    lbl.image=img1
    
    
    img=""
    

#save

def save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender 
    except:
        messagebox.showerror("error","Select Gender")
    
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=State.get()
    father_name=F_Name.get()
    mother_name=M_Name.get()
    F1=Father_Occupation.get()
    F2=Name.get()
    
    
    if N=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or father_name=="" or mother_name=="" or F1=="" or  F2=="":
       messagebox.showerror("error","Few Data is missing")
    else:
       file=openpyxl.load_workbook('student_data.xlsx')
       sheet=file.active
       sheet.cell(column=1,row=sheet.max_row+1,value=R1)
       sheet.cell(column=2,row=sheet.max_row,value=N1)
       sheet.cell(column=3,row=sheet.max_row,value=C1)
       sheet.cell(column=4,row=sheet.max_row,value=G1)
       sheet.cell(column=5,row=sheet.max_row,value=D2)
       sheet.cell(column=6,row=sheet.max_row,value=D1)
       sheet.cell(column=7,row=sheet.max_row,value=Re1)
       sheet.cell(column=8,row=sheet.max_row,value=S1)
       sheet.cell(column=9,row=sheet.max_row,value=father_name)
       sheet.cell(column=10,row=sheet.max_row,value=mother_name)
       sheet.cell(column=11,row=sheet.max_row,value=F1)

       file.save(r"student_data.xlsx")
       
       
       
       
       try :
           img.save("Student Images/"+str(R1)+".jpg")
       except:
           messagebox.showinfo("info","Profile Picture is not available!!!!")
    messagebox.showinfo("info","successfully data entered!!!!")
    Clear()
    registration_no()
    
    
    
    
    
    
    
#exit

def Exit():
    root.destroy()




    
#gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
       
    else:
        gender="Female"
          
    
    
#Top Frame E
Label(root,text="Email: kmtammalwad@gmail.com", width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)    
Label(root,text="STUDENT  DETAILS", width=18,height=2,bg="#c36464",fg='#fff',font='arial 20 ').pack(side=TOP,fill=X)    

#search box
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
imageicon3=PhotoImage(file="Images/search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg="#68ddfa",font="arial 13 bold",command=search)
Srch.place(x=1060,y=70)
imageicon4=PhotoImage(file="Images/Layer4.png")
Update_button=Button(root,image=imageicon4,bg="orange",command=Update)
Update_button.place(x=110,y=51)


#registration and date
Label(root,text="Registration No:",font="arial 12",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 12",fg=framebg,bg=background).place(x=500,y=150)
Registration=IntVar()
Date=StringVar()
reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)
registration_no()

#date
today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)
Date.set(d1)

#student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name :",font="arial 12",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth :",font="arial 12",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender :",font="arial 12",bg=framebg,fg=framefg).place(x=30,y=150)


Label(obj,text="Class :",font="arial 12",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion :",font="arial 12",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="State :",font="arial 12",bg=framebg,fg=framefg).place(x=500,y=150)
Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=15,font="arial 20")
name_entry.place(x=160,y=50)


DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=15,font="arial 20")
dob_entry.place(x=160,y=100)


radio=IntVar()
R1 = Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)
R2 = Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=150,y=180)
#regilion
Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=15,font="arial 20")
religion_entry.place(x=630,y=100)


#state
State=StringVar()
state_entry=Entry(obj,textvariable=State,width=15,font="arial 20")
state_entry.place(x=630,y=150)

#class
Class= Combobox(obj,values=['CIVIL','ENTC','MECHANICAL','COMPUTER','CHEMICAL'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")



#company details 
obj=LabelFrame(root,text="Company Details",font=20,bd=4,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj.place(x=30,y=470)
Label(obj,text="Company Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Internship Duration:",font="arial 13",bg=framebg,fg=framefg).place(x=0,y=100)



F_Name=StringVar()
f_entry=Entry(obj,textvariable=F_Name,width=15,font="arial 20")
f_entry.place(x=160,y=50)

Father_Occupation=StringVar()
fo_entry=Entry(obj,textvariable=Father_Occupation,width=15,font="arial 20")
fo_entry.place(x=160,y=100)




Label(obj,text="Company Address:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)



M_Name=StringVar()
M_entry=Entry(obj,textvariable=M_Name,width=15,font="arial 20")
M_entry.place(x=660,y=50)


#image
f=Frame(root,bd=3,bg="white",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=200)

img=PhotoImage(file="Images/upload.png")
lbl=Label(f,bg="white",image=img)
lbl.place(x=0,y=0)
#buttons
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=390)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightblue",command=save)
saveButton.place(x=1000,y=460)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="gray",command=Exit).place(x=1000,y=600)

root.mainloop()
